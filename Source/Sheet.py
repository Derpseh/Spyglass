# Whee, lots of imports!
import urllib
import gzip
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.styles.colors import RED, GREEN
import xml.etree.ElementTree as ET
import math
from datetime import datetime
import os
import urllib2

# Gotta set my headers
UAgent = str(raw_input('Nation Name: '))
headers = { 'User-Agent' : 'Spyglass. Currently in use by Panzer Vier (Authenticating). Devved by Panzer Vier > valkynora@gmail.com'}
try:
	testReq = urllib2.Request('http://www.nationstates.net/cgi-bin/api.cgi?nation=' + UAgent.replace(' ', '_') + "&q=influence")
	testhtml = urllib2.urlopen(testReq).read()
	headers = { 'User-Agent' : 'Spyglass. Currently in use by ' + UAgent + '. Devved by Panzer Vier > valkynora@gmail.com'}
except:
	print "Nation not found. Be sure to input the name of a nation that actually exists"
	quit()
print "Pulling Data Dump..."
# Pulling a list of regions that are founderless and non-passworded. Eventually, we'll go through and highlight those on the sheet
req = urllib2.Request('http://www.nationstates.net/cgi-bin/api.cgi?q=regionsbytag;tags=founderless,-password', None, headers)
html = urllib2.urlopen(req).read()
# Grabbing the data dump and saving
urllib.urlretrieve ('http://www.nationstates.net/pages/regions.xml.gz', 'regions.xml.gz')
# Gotta get the date so we can attach it to the filename afterwards.So, I suppose, run the thing sometime close to midnight, and you might end up with the wrong date.
now = datetime.now()
YMD = '%s-%s-%s' % (now.year, now.month, now.day)
redFill = PatternFill(start_color = RED, end_color = RED, fill_type = 'solid')
greenFill = PatternFill(start_color = GREEN, end_color = GREEN, fill_type = 'solid')

if os.path.isfile('UpdTime'):
	with open('UpdTime', 'r') as UpdFile:
		UpdTime = UpdFile.read()
else:
	with open('UpdTime', 'w') as UpdFile:
		UpdTime = "3741, 5412"
		UpdFile.write("3811, 5428")
UpdTime = UpdTime.split(', ')
MinorTime = int(UpdTime[0])
MajorTime = int(UpdTime[1])

# Un-gzipping
with gzip.open('regions.xml.gz', 'rb') as infile:
    with open('regions.xml', 'w') as outfile:
        for line in infile:
            outfile.write(str(line))

# Opening up my virtual sheet. Gotta find a better name for it, sometime. The pink tab colour's pretty sweet, tho.
wb = Workbook()
ws = wb.active
ws.title = "Spyglass Timesheet"
ws.sheet_properties.tabColor = "FFB1B1"

RegionList = []
RegionURLList = []
NumNationList = []
DelVoteList = []

# Sanitize our founderless regions list a wee bit, 'cause at the moment, it's xml, and xml is gross.
print "Processing data..."
root = ET.fromstring(html)
for EVENT in root.iter('REGIONS'):
    UnfoundedString = EVENT.text
UnfoundedList = UnfoundedString.split(',')
# Not entirely sure why I don't just keep this within the program, rather than saving and reloading it, but oh wells.
with open('regions.xml', 'r') as myfile:
    regions=myfile.read()
# Pulling, in order, region names, converting to a region url, number of nations in that region, and voting power that delegate has.
root = ET.fromstring(regions)
for EVENT in root.iter('NAME'):
    RegionList = RegionList + [EVENT.text]
    UrlString = '=HYPERLINK("http://www.nationstates.net/region=' + EVENT.text + '")'
    # UrlString.replace(' ', '_')
    RegionURLList = RegionURLList + [UrlString.replace(' ', '_')]
for EVENT in root.iter('NUMNATIONS'):
    NumNationList = NumNationList + [int(EVENT.text)]
for EVENT in root.iter('DELEGATEVOTES'):
    DelVoteList = DelVoteList + [int(EVENT.text)]
# Grabbing the cumulative number of nations that've updated by the time a region has.
CumulNationList = []
for a in NumNationList:
    if len(CumulNationList) == 0:
        CumulNationList.extend([int(a)])
    else:
        CumulNationList.extend([int(a) + CumulNationList[-1]])

# Still need to get this to auto-calibrate. Subject to change in the near future, but the times are a rough approximation anyway, so...
CumulNations = CumulNationList[-1]
MinorNatTime = float(MinorTime) / CumulNations
MajorNatTime = float(MajorTime) / CumulNations
MinTime = []
MajTime = []
# Getting the approximate major/minor update times.
for a in CumulNationList:
    temptime = round(a * MinorNatTime)
    temptime2 = round(a * MajorNatTime)
    tempsecs = int(temptime % 60)
    tempmins = int(math.floor(temptime / 60) % 60)
    temphours = int(math.floor(temptime / 3600))
    tempsecs2 = int(temptime2 % 60)
    tempmins2 = int(math.floor(temptime2 / 60) % 60)
    temphours2 = int(math.floor(temptime2 / 3600))
    MinTime.extend(['%s:%s:%s' % (temphours, tempmins, tempsecs)])
    MajTime.extend(['%s:%s:%s' % (temphours2, tempmins2, tempsecs2)])

#Splashing some headers and stuff onto the spreadsheet for legibility purposes!
ws['A1'].value = 'Regions'
ws['B1'].value = 'Region Link'
ws['C1'].value = '# Nations'
ws['D1'].value = 'Tot. Nations'
ws['E1'].value = 'Minor Upd.'
ws['F1'].value = 'Major Upd.'
ws['G1'].value = 'Del. Votes'
ws['H1'].value = 'Del. Endos'
ws['J1'].value = 'World '
ws['K1'].value = 'Data'
ws['J2'].value = 'Nations'
ws['J3'].value = 'Last Major'
ws['J4'].value = 'Secs/Nation'
ws['J5'].value = 'Nations/Sec'
ws['J6'].value = 'Last Minor'
ws['J7'].value = 'Secs/Nation'
ws['J8'].value = 'Nations/Sec'
ws['K2'].value = CumulNations
ws['K3'].value = MajorTime
ws['K4'].value = MajorNatTime
ws['K5'].value = 1 / MajorNatTime
ws['K6'].value = MinorTime
ws['K7'].value = MinorNatTime
ws['K8'].value = 1 / MinorNatTime

# There's probably a better way of doing this, but my coding skills are dubious :^)
# Anyways, actually pasting the information from our various lists into the spreadsheet.
counter = 0

for a in RegionList:
# If a region's founderless, highlight it for easy reference. Add a tilde, 'cause my spreadsheet program doesn't do filtering by colour
    if a in UnfoundedList:
        ws.cell(row = counter + 2, column = 1).fill = greenFill
        ws.cell(row = counter + 2, column = 2).fill = greenFill
        b = a + '~'
    else:
        b = a
    ws.cell(row = counter + 2, column = 1).value = b
    ws.cell(row = counter + 2, column = 2).value = RegionURLList[counter]
    ws.cell(row = counter + 2, column = 3).value = NumNationList[counter]
    ws.cell(row = counter + 2, column = 4).value = CumulNationList[counter]
    ws.cell(row = counter + 2, column = 5).value = MinTime[counter]
    ws.cell(row = counter + 2, column = 5).alignment = Alignment(horizontal="right")
    ws.cell(row = counter + 2, column = 6).value = MajTime[counter]
    ws.cell(row = counter + 2, column = 6).alignment = Alignment(horizontal="right")
    ws.cell(row = counter + 2, column = 7).value = DelVoteList[counter]
    ws.cell(row = counter + 2, column = 8).value = DelVoteList[counter] - 1
# Highlight delegate-less regions. They're good for tagging, or whatever~
    if DelVoteList[counter] == 0:
        ws.cell(row = counter + 2, column = 8).fill = redFill
    counter += 1
# You know those situations where you can't quite get code to work, and kinda fumble around until you find something that does?
# I'm 90% sure this isn't the way to do it, but I couldn't get it working otherwise.
# Anyways, setting the region name column's width, so that it doesn't cut everything off.
sheet = wb.get_sheet_by_name('Spyglass Timesheet')
sheet.column_dimensions['A'].width = 45
sheet['J1'].alignment = Alignment(horizontal="right")


# Really should just name the sheets 'Derps is amazing in every conceivable way'. Would be some free ego-massage.
print "Saving Sheet..."
wb.save('SpyglassSheet' + YMD + '.xlsx')
# Deleting the dump, 'cause not needed anymore. Let's keep things vaguely clean, neh?
print "Cleaning up..."
os.remove('regions.xml.gz')
os.remove('regions.xml')

#!/usr/bin/env python3
from requests import get
from requests.exceptions import HTTPError
import gzip
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.styles.colors import Color
from datetime import datetime
from pathlib import Path
from sys import argv
from typing import List
from xml.etree import ElementTree

# UPDATE THIS EVERY TIME A NEW RELEASE IS PACKAGED
VERSION = "2.0.1"

# Spyglass
# Source code by Devi aka Panzer Vier
# Modifications made by Khronion (KH)
# Ported to Python 3 with additional modifications by Zizou (Ziz)
# Yay more modifications (Aav)

log_path = "debug.log"

RED = Color(rgb="FFFF0000")
GREEN = Color(rgb="FF00FF00")
YELLOW = Color(rgb="FFFFFF00")


# Method for writing a debug log
def write_log(to_write: str) -> None:
    """
    Writes a string to the debug log.
    :param to_write: str
    :return: None
    """
    if log:
        with open(log_path, "a") as out:
            out.write(datetime.now().strftime(f"[%Y-%m-%d %H:%M:%S] {to_write}\n"))
    else:
        pass


# Method for getting user input
def query(args: str, options: List[str]) -> str:
    """
    Gets user input from the command line and checks it against possible options.
    :param args: str
    :param options: List[str]
    :return: str
    """
    while True:
        response = input(args)
        if response in options:
            return response


# Method for downloading data dump, and saving it to disk
def download_dump() -> None:
    """
    Downloads the most recent daily dump from NS.
    :return: None
    """
    dump_request = get(
        "https://www.nationstates.net/pages/regions.xml.gz", stream=True
    )
    with open("regions.xml.gz", "wb") as data_dump:
        for chunk in dump_request.iter_content(chunk_size=16 * 1024):
            data_dump.write(chunk)


# Parse arguments, if any...

# Show help message and terminate
if "-h" in argv or "--help" in argv:
    print(f"Spyglass {VERSION}: Generate NationStates region update timesheets.\n")
    print("Developed by Panzer Vier, with additions by Khronion, Zizou, and Aav\n")
    print(f"usage: {argv[0]} [-h] [-n NATION] [-o OUTFILE] [-s | -l PATH]\n")
    print(
        """Optional arguments:
     -h           Show this help message and exit.
     -n NATION    Specify Nation to identify user by. In order to comply with
                  NationStates API rules, this must be the user's nation. Use
                  underscores instead of spaces.
     -o OUTFILE   File to output the generated timesheet in XLSX format to.
     -s           Suppress creating a debug log file. Log files are written to
                  the current working directory.
     -l PATH      Write debug log to specified path.
     -m           Generate a minimized sheet without WFEs and embassies
    """
    )
    print(
        """If run without arguments, Spyglass runs in interactive mode and outputs to its
working directory."""
    )
    raise SystemExit(1)

interactive = True
process_embassies = True
log = True

SpeedOverride = False
MinorTime = 3550
MajorTime = 5350

YMD = f"{datetime.now().year}-{datetime.now().month}-{datetime.now().day}"

# Set nation name
if "-n" in argv:
    interactive = False
    UAgent = argv[argv.index("-n") + 1]
else:
    print(f"Spyglass {VERSION}: Generate NationStates region update timesheets.")
    UAgent = input("Nation Name: ")
    filename = f"SpyglassSheet{YMD}.xlsx"

    if query("Include region embassies? (y/n, defaults to y) ", ["y", "n", ""]) == "n":
        process_embassies = False

    # Ziz: Update lengths are now 1.5hrs and 2.5hrs for minor and major respectively
    if (
            query(
                "Do you want to manually specify update lengths? (y/n, defaults to n) ",
                ["y", "n", ""],
            )
            == "y"
    ):
        try:
            MinorTime = int(input("Minor Time, seconds (3550): "))
        except ValueError:
            MinorTime = 3550
        try:
            MajorTime = int(input("Major Time, seconds (5350): "))
        except ValueError:
            MajorTime = 5350
        SpeedOverride = True

# Set output filename
if "-o" in argv:
    filename = argv[argv.index("-o") + 1]
else:
    filename = f"SpyglassSheet{YMD}.xlsx"

# Enable debug log
if "-s" in argv:
    log = False

if "-m" in argv:
    process_embassies = False
else:
    if "-l" in argv:
        log_path = argv[argv.index("-l") + 1]
    starting_args = " ".join(argv[1:])
    write_log(f"INFO Spyglass started with arguments: {starting_args}")
    write_log(f"INFO User Nation: {UAgent}")
    write_log(f"INFO Out File: {filename}")

# Set headers as required by NS TOS
headers = {
    "User-Agent": f"Spyglass/{VERSION} (github: https://github.com/Derpseh/Spyglass ; user:{UAgent}; Authenticating)"
}

# Verify specified nation is valid -- terminate if not
try:
    params = {"nation": UAgent, "q": "influence"}
    testreq = get(
        "https://www.nationstates.net/cgi-bin/api.cgi", headers=headers, params=params
    )
    testreq.raise_for_status()
    headers = {
        "User-Agent": f"Spyglass/{VERSION} (github: https://github.com/Derpseh/Spyglass ; user:{UAgent})"
    }
except HTTPError:
    print(
        "Nation not found. Be sure to input the name of a nation that actually exists."
    )
    write_log(f"ERR  {UAgent} is not a valid nation. Terminating.")
    raise SystemExit(1)

write_log(f"INFO Minor length: {MinorTime}")
write_log(f"INFO Major length: {MajorTime}")

# Pulling a list of regions that are founderless and non-passworded. Eventually, we'll go through and highlight those
# on the sheet

write_log("INFO Searching for data dump...")

# Ziz: If a data dump is detected in the current directory, ask if user wants to re-download latest dump
# Ziz: Otherwise just download the latest data dump if nothing is detected
dump_path = Path("./regions.xml.gz")
if interactive:
    if dump_path.exists() and dump_path.is_file():
        if (
                query(
                    "Existing data dump found. Do you want to re-download the latest dump? (y/n, defaults to y) ",
                    ["y", "n", ""],
                )
                == "y"
        ):
            write_log("INFO Found data dump, but re-downloading the latest..")
            print("Pulling data dump...")
            download_dump()
            write_log("INFO Download complete!")
        else:
            write_log("INFO Using data dump already present...")
            print("Using current dump...")
    else:
        write_log("INFO No existing data dump found, downloading latest...")
        print("No existing data dump found. Pulling data dump...")
        download_dump()
        write_log("INFO Download complete!")
else:
    write_log("INFO running in non-interactive mode, downloading data dump...")
    print("Pulling data dump...")
    download_dump()
    write_log("INFO Download complete!")

redFill = PatternFill(start_color=RED, end_color=RED, fill_type="solid")
greenFill = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
yellowFill = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")

# Un-gzipping
# Ziz: Now we can just decompress the dump and hand it to the parser without writing to disk
with gzip.open("regions.xml.gz", "rb") as dump:
    regions = dump.read()

# Opening up my virtual sheet. Gotta find a better name for it, sometime. The pink tab colour's pretty sweet, tho.
wb = Workbook()
ws = wb.active
ws.title = "Spyglass Timesheet"
ws.sheet_properties.tabColor = "FFB1B1"

RegionList = list()
RegionURLList = list()
RegionWFEList = list()
RegionEmbassyList = list()
NumNationList = list()
DelVoteList = list()
ExecList = list()
MajorList = list()

# Sanitize our founderless regions list a wee bit, 'cause at the moment, it's xml, and xml is gross.
print("Processing data...")
# Total number of queries is low, so rate limit is unnecessary
req = get(
    "https://www.nationstates.net/cgi-bin/api.cgi?q=regionsbytag;tags=-password",
    headers=headers,
).text
req2 = get(
    "https://www.nationstates.net/cgi-bin/api.cgi?q=regionsbytag;tags=founderless",
    headers=headers,
).text

unfounded_list = ElementTree.fromstring(req2).find("REGIONS").text.split(",")
pwless_list = ElementTree.fromstring(req).find("REGIONS").text.split(",")

# Pulling, in order, region names, converting to a region url, number of nations in that region, and voting power that
# delegate has.

# Ziz: Data dump soup, mmm... almost as tasty as people!
# Aav: Refactored old code to use zip. I did leave Ziz's cursed comment though
data = ElementTree.fromstring(regions)
region_list = data.findall("REGION")
names = [region.find("NAME") for region in region_list]
num_nations = [region.find("NUMNATIONS") for region in region_list]
delvotes = [region.find("DELEGATEVOTES") for region in region_list]
delauth = [region.find("DELEGATEAUTH") for region in region_list]
for name, nation_count, del_votes, auth in zip(
        names, num_nations, delvotes, delauth
):
    RegionList.append(name.text)
    UrlString = f'=HYPERLINK("https://www.nationstates.net/region={name.text}")'
    RegionURLList.append(UrlString.replace(" ", "_"))
    NumNationList.append(int(nation_count.text))
    DelVoteList.append(int(del_votes.text))
    AuthString = auth.text
    if AuthString[0] == "X":
        ExecList.append(True)
    else:
        ExecList.append(False)

# KH: pull major times from daily dump
# Aav: Refactored into listcomp 3/17/2022
MajorList = [int(d.find("LASTUPDATE").text) for d in region_list]

# KH: gather WFE info
for wfe in [d.find("FACTBOOK") for d in region_list]:
    text = wfe.text
    try:
        if text[0] in ["=", "+", "-", "@"]:
            text = f"'{text}"  # IMPORTANT: prevent excel from parsing WFEs as formulas
        RegionWFEList.append(text)
    except TypeError:  # no WFE
        RegionWFEList.append("")

for region_embassies in [d.find("EMBASSIES") for d in region_list]:
    embassies = list()
    if process_embassies:
        for embassy in region_embassies.findall("EMBASSY"):
            embassies.append(embassy.text)
    RegionEmbassyList.append(",".join(embassies))

# Determine the total duration in seconds of minor and major
if not SpeedOverride:
    major = MajorList[-1] - MajorList[0]

# ...unless we're overriding it
else:
    major = int(MajorTime)

# Grabbing the cumulative number of nations that've updated by the time a region has.
# The first entry is zero because time calculations need to reflect the start of region update, not the end
CumulNationList = [0]
for a in NumNationList:
    CumulNationList.append(int(a) + CumulNationList[-1])

# Calculate speed based on total population
CumulNations = CumulNationList[-1]
MinorNatTime = int(MinorTime) / CumulNations
MajorNatTime = major / CumulNations
MinTime = list()
MajTime = list()

# Getting the approximate major/minor update times.
for a in CumulNationList:
    temptime = int(a * MinorNatTime)
    tempsecs = temptime % 60
    tempmins = int((temptime // 60) % 60)
    temphours = int(temptime // 3600)
    MinTime.append(f"{temphours}:{tempmins}:{tempsecs}")

# If user specifies update length, use special handling.
if SpeedOverride:
    for a in CumulNationList:
        temptime = int(a * MajorNatTime)
        tempsecs = temptime % 60
        tempmins = int((temptime // 60) % 60)
        temphours = int(temptime // 3600)
        MajTime.append(f"{temphours}:{tempmins}:{tempsecs}")
else:
    for a in MajorList:
        temptime = a - MajorList[0]
        tempsecs = temptime % 60
        tempmins = int((temptime // 60) % 60)
        temphours = int(temptime // 3600)
        MajTime.append(f"{temphours}:{tempmins}:{tempsecs}")

# Splashing some headers and stuff onto the spreadsheet for legibility purposes!
ws["A1"].value = "Regions"
ws["B1"].value = "Region Link"
ws["C1"].value = "# Nations"
ws["D1"].value = "Tot. Nations"
ws["E1"].value = "Minor Upd. (est)"
if SpeedOverride is True:
    ws["F1"].value = "Major Upd. (est)"
else:
    ws["F1"].value = "Major Upd. (true)"
ws["G1"].value = "Del. Votes"
ws["H1"].value = "Del. Endos"
if process_embassies:
    ws["I1"].value = "Embassies"
ws["J1"].value = "WFE"

ws["L1"].value = "World "
ws["M1"].value = "Data"
ws["L2"].value = "Nations"
ws["L3"].value = "Last Major"
ws["L4"].value = "Secs/Nation"
ws["L5"].value = "Nations/Sec"
ws["L6"].value = "Last Minor"
ws["L7"].value = "Secs/Nation"
ws["L8"].value = "Nations/Sec"
ws["L10"].value = "Spyglass Version"
ws["L11"].value = "Date Generated"
ws["M2"].value = CumulNations
ws["M3"].value = major
ws["M4"].value = major / CumulNations
ws["M5"].value = 1 / (major / CumulNations)
ws["M6"].value = MinorTime
ws["M7"].value = MinorNatTime
ws["M8"].value = 1 / MinorNatTime
ws["M10"].value = VERSION
ws["M11"].value = YMD

# There's probably a better way of doing this, but my coding skills are dubious :^)
# Anyways, actually pasting the information from our various lists into the spreadsheet.

# Aav: Change to using enumerate instead of an external counter
for counter, a in enumerate(RegionList):
    # If a region's founderless, highlight it for easy reference. Add a tilde, 'cause my spreadsheet program doesn't
    # do filtering by colour

    b = a
    # KH: ~ indicates hittable
    # KH: yellow = passwordless and exec delegate
    if a in pwless_list and ExecList[counter]:
        ws.cell(row=counter + 2, column=1).fill = yellowFill
        ws.cell(row=counter + 2, column=2).fill = yellowFill
        b = f"{a}~"
    # KH: green = founderless and passwordless
    if a in unfounded_list and a in pwless_list:
        ws.cell(row=counter + 2, column=1).fill = greenFill
        ws.cell(row=counter + 2, column=2).fill = greenFill
        b = f"{a}~"
    # KH: red = passwordless
    if a not in pwless_list:
        ws.cell(row=counter + 2, column=1).fill = redFill
        ws.cell(row=counter + 2, column=2).fill = redFill
        b = f"{a}*"
    ws.cell(row=counter + 2, column=1).value = b
    ws.cell(row=counter + 2, column=2).value = RegionURLList[counter]
    ws.cell(row=counter + 2, column=3).value = NumNationList[counter]
    ws.cell(row=counter + 2, column=4).value = CumulNationList[counter]
    ws.cell(row=counter + 2, column=5).value = MinTime[counter]
    ws.cell(row=counter + 2, column=5).alignment = Alignment(horizontal="right")
    ws.cell(row=counter + 2, column=6).value = MajTime[counter]
    ws.cell(row=counter + 2, column=6).alignment = Alignment(horizontal="right")
    ws.cell(row=counter + 2, column=7).value = DelVoteList[counter]
    ws.cell(row=counter + 2, column=8).value = DelVoteList[counter] - 1
    ws.cell(row=counter + 2, column=9).value = RegionEmbassyList[counter]
    ws.cell(row=counter + 2, column=10).value = RegionWFEList[counter]
    ws.cell(row=counter + 2, column=11).value = " "

    # Highlight delegate-less regions. They're good for tagging, or whatever~
    if DelVoteList[counter] == 0:
        ws.cell(row=counter + 2, column=8).fill = redFill

# You know those situations where you can't quite get code to work, and kinda fumble around until you find something
# that does?
#
# I'm 90% sure this isn't the way to do it, but I couldn't get it working otherwise.
# Anyways, setting the region name column's width, so that it doesn't cut everything off.
sheet = wb["Spyglass Timesheet"]
sheet.column_dimensions["A"].width = 45
sheet["J1"].alignment = Alignment(horizontal="right")

write_log("INFO Done processing data! Saving sheet.")

# Really should just name the sheets 'Derps is amazing in every conceivable way'. Would be some free ego-massage.
print("Saving sheet...")
wb.save(filename)
# Ziz: Don't delete the data dump, since it can be reused if it's recent enough

write_log(f"INFO Successfully saved to {filename}")
write_log(f"INFO Spyglass run complete. Exiting...")

raise SystemExit(0)

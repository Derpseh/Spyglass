from requests import Session
from requests.exceptions import HTTPError
import gzip
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.styles.colors import Color
from datetime import datetime
from pathlib import Path
import argparse
from typing import List, Optional
from lxml import etree
import logging

# Spyglass
# Source code by Devi aka Panzer Vier
# Modifications made by Khronion (KH)
# Ported to Python 3 with additional modifications by Zizou (Ziz)
# Yay more modifications & V3 rewrite (Aav)

# UPDATE THIS EVERY TIME A NEW RELEASE IS PACKAGED
VERSION = "3.0.1"

# Color declarations because OpenPyXL removed them :^)
RED = Color(rgb="FFFF0000")
GREEN = Color(rgb="FF00FF00")
YELLOW = Color(rgb="FFFFFF00")

# PatternFill declarations
redFill = PatternFill(start_color=RED, end_color=RED, fill_type="solid")
greenFill = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
yellowFill = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")

# Set up our default settings
interactive = True
minimize = False
nation = ""
SpeedOverride = False

# Set up logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    filename="spyglass.log",
    filemode="w",
)
logger = logging.getLogger(__name__)


# Method for getting user input
def query(inargs: str, options: List[str]) -> str:
    """
    Gets user input from the command line and checks it against possible options.
    :param inargs: str
    :param options: List[str]
    :return: str
    """
    while True:
        response = input(inargs)
        if response in options:
            return response


# Method for downloading data dump, and saving it to disk
def download_dump(rsession: Session) -> None:
    """
    Downloads the most recent daily dump from NS.
    :return: None
    """
    dump_request = rsession.get(
        "https://www.nationstates.net/pages/regions.xml.gz", stream=True
    )
    with open("regions.xml.gz", "wb") as data_dump:
        for chunk in dump_request.iter_content(chunk_size=16 * 1024):
            data_dump.write(chunk)


def sanitize(string: str) -> Optional[str]:
    """
    Prevents Excel from interpreting strings as formulas.
    Thanks Khronion for the original catch & fix.
    -- Aav
    :param string:
    :return:
    """
    # noinspection PyBroadException
    try:
        if string[0] in ["=", "+", "-", "@"]:
            return f"'{string}"
        else:
            return string

    except Exception:
        return None


# Set up command line arguments
parser = argparse.ArgumentParser(
    prog="Spyglass",
    description=f"Spyglass {VERSION}: Generate NationStates region update timesheets.\nDeveloped by Panzer Vier, "
    f"with additions by Khronion, Zizou, and Aav",
    add_help=True,
)
parser.add_argument(
    "-n",
    "--nation",
    help="Nation to use for authentication. Use underscores instead of spaces.",
    required=False,
)
parser.add_argument(
    "-o",
    "--outfile",
    help="Name of the output file. Defaults to 'spyglass.xlsx'.",
    default=f"spyglass{datetime.now().year}-{datetime.now().month}-{datetime.now().day}.xlsx",
    required=False,
)
parser.add_argument(
    "-m",
    "--minimize",
    help="Generate a minimized sheet without WFEs and embassies",
    action="store_true",
    required=False,
)
parser.add_argument(
    "-s",
    "--suppress",
    help="Suppresses the creation of a debug log file",
    action="store_true",
    required=False,
)
parser.add_argument(
    "--minor",
    help="The length of minor update in seconds.",
    default=3550,
    required=False,
)
parser.add_argument(
    "--major",
    help="The length of major update in seconds.",
    default=5350,
    required=False,
)
parser.add_argument(
    "-d",
    "--dump",
    help="Do not download the latest data dump. Use the one in the current directory.",
    action="store_true",
    required=False,
)
parser.add_argument(
    "-p",
    "--path",
    help="Path to the data dump. Defaults to 'regions.xml.gz'.",
    default="./regions.xml.gz",
    required=False,
)

args = parser.parse_args()

if args.suppress:
    logger.disabled = True  # Disable logging

if args.minimize:
    minimize = True  # Disable WFEs and embassies

if args.major != 5350:
    SpeedOverride = True  # Override the default update times

logger.info(f"Spyglass {VERSION} started")
logger.info(f"Starting with arguments: {args}")

# If a nation is supplied, we run in headless mode
if args.nation:
    interactive = False
else:
    # We need to get a nation to set as the useragent for the session
    nation = input("Please enter your nation name: ").lower().replace(" ", "_")
    logger.info(f"User entered nation: {nation}")

# Construct our requests session now that we have an useragent
session = Session()
session.headers.update(
    {
        "User-Agent": f"Spyglass/{VERSION} (github: https://github.com/Derpseh/Spyglass; user:{nation}; Authenticating)"
    }
)
logger.info("Session constructed")

# Ensure that this nation actually exists
try:
    req = session.get(
        f"https://www.nationstates.net/cgi-bin/api.cgi?nation={nation}&q=region"
    )
    req.raise_for_status()
except HTTPError:
    logger.error("Nation does not exist.")
    print("The provided nation does not exist... terminating.")
    raise SystemExit("Authentication failure.")

logger.info("Nation exists, proceeding.")
session.headers.update(
    {
        "User-Agent": f"Spyglass/{VERSION} (github: https://github.com/Derpseh/Spyglass; user:{nation})"
    }
)
logger.info("Updated useragent.")
logger.info("Checking if there is a data dump downloaded.")
# Ziz: If a data dump is detected in the current directory, ask if user wants to re-download latest dump
# Ziz: Otherwise just download the latest data dump if nothing is detected

dump_path = Path(args.path)
if interactive:
    if dump_path.exists() and dump_path.is_file():
        if (
            query(
                "Existing data dump found. Do you want to re-download the latest dump? (y/n) ",
                ["y", "n", ""],
            )
            == "y"
        ):
            logger.info("Found data dump, but re-downloading the latest..")
            print("Pulling data dump...")
            download_dump(session)
            logger.info("Download complete!")
        else:
            logger.info("Using data dump already present...")
            print("Using current dump...")
    else:
        logger.info("No existing data dump found, downloading latest...")
        print("No existing data dump found. Pulling data dump...")
        download_dump(session)
        logger.info("Download complete!")
else:
    if dump_path.exists() and dump_path.is_file():
        if args.dump:
            logger.info("Using data dump already present...")
        else:
            download_dump(session)
            logger.info("Download complete!")
    else:
        logger.info("No existing data dump found, downloading latest...")
        download_dump(session)
        logger.info("Download complete!")

# Get the lists of founderless and passwordless regions
logger.info("Getting founderless regions...")
fless = session.get(
    "https://www.nationstates.net/cgi-bin/api.cgi?q=regionsbytag;tags=founderless"
).text.split(",")
logger.info("Getting passwordless regions...")
pless = session.get(
    "https://www.nationstates.net/cgi-bin/api.cgi?q=regionsbytag;tags=-password"
).text.split(",")

# Open up the data dump and parse it
logger.info("Parsing data dump...")
dump = gzip.open(args.path, "rb").read()

# Ziz: Data dump soup, mmm... almost as tasty as people!
# Aav, at some point in the past: Refactored old code to use zip. I did leave Ziz's cursed comment though
# Aav 3/20/23: It's 2023, and we've finally adopted lxml... it's only been out since 2005.
data = etree.fromstring(dump)
region_list = data.findall("REGION")
logger.info("Data dump parsed successfully!")

# Now that we created a data list, purge the data dump from memory
del dump
del data
logger.info("Data dump purged from memory.")

# Happiness ensues for low-memory systems!
regions = []
names = [region.find("NAME") for region in region_list]
num_nations = [region.find("NUMNATIONS") for region in region_list]
delvotes = [region.find("DELEGATEVOTES") for region in region_list]
delauth = [region.find("DELEGATEAUTH") for region in region_list]
MajorList = [int(d.find("LASTUPDATE").text) for d in region_list]
wfelist = [d.find("FACTBOOK") for d in region_list]
logger.info("List parsing complete!")

# Aav: Finding embassies outside of that zip loop and then referencing inside is probably faster than searching
# the entire dump every. single. time.

embassies = {}
for region in region_list:
    name = region.find("NAME").text
    l = []
    for embassy in region.find("EMBASSIES").findall("EMBASSY"):
        l.append(embassy.text)
    embassies[name] = l

logger.info("Embassy parsing complete!")

for name, nation_count, del_votes, auth, time, wfe in zip(
    names, num_nations, delvotes, delauth, MajorList, wfelist
):
    _ = {
        "name": name.text,
        "url": f'=HYPERLINK("https://www.nationstates.net/region={name.text}")',
        "num_nations": int(nation_count.text),
        "del_votes": int(del_votes.text),
        "exec": auth.text[0] == "X",
        "last_update": time,
        "wfe": sanitize(wfe.text) if not minimize else None,
        "embassies": embassies[name.text]
        if not minimize or not len(embassies[name.text])
        else None,
    }
    regions.append(_)
logger.info("Region dictionary created!")

# Get rid of that parsed XML
# Gotta go FASSSSSSSTTTTTTT
del region_list

major = int(args.major)
minor = int(args.minor)

if not SpeedOverride:
    # Calculate major based off of the daily dump
    last_region = regions[-1]
    first_region = regions[0]
    major = last_region["last_update"] - first_region["last_update"]
    logger.info(f"Major calculated as {major}")
    logger.info(f"Minor set as {minor}")
else:
    logger.info("Speed override enabled, skipping major calculation.")
    logger.info(f"Major set to {major}")
    logger.info(f"Minor set to {minor}")

CumuNatList = [
    0
]  # Per Devi, the first number needs to be zero so that the times reflect the start of update,
# not the end

for region in regions:
    CumuNatList.append(CumuNatList[-1] + region["num_nations"])
    region.update({"cumu_nations": CumuNatList[-1] + region["num_nations"]})
logger.info("Cumulative nation list created!")
logger.info(f"Total number of nations: {CumuNatList[-1]}")

# Calculate the time per nation for major and minor
major_per_nation = major / CumuNatList[-1]
minor_per_nation = minor / CumuNatList[-1]

for region in regions:
    nats = region["cumu_nations"]
    tempminor = int(nats * minor_per_nation)
    tempmajor = int(nats * major_per_nation)
    # Convert seconds to hours, minutes, seconds
    # ti_x = tempminor, tm_x = tempmajor
    ti_s = tempminor % 60
    ti_m = int((tempminor // 60) % 60)
    ti_h = int((tempminor // 3600))
    region.update({"minor": f"{ti_h}:{ti_m}:{ti_s}"})
    tm_s = tempmajor % 60
    tm_m = int((tempmajor // 60) % 60)
    tm_h = int((tempmajor // 3600))
    region.update({"major": f"{tm_h}:{tm_m}:{tm_s}"})
logger.info("Speeds calculated")

# Create the spreadsheet
wb = Workbook()
ws = wb.active
ws.title = "Spyglass Timesheet"
ws.sheet_properties.tabColor = "FFB1B1"
logger.info("Spreadsheet created!")

# Splashing some headers and stuff onto the spreadsheet for legibility purposes!
ws["A1"].value = "Regions"
ws["B1"].value = "Region Link"
ws["C1"].value = "# Nations"
ws["D1"].value = "Tot. Nations"
ws["E1"].value = "Minor Upd. (est)"
if SpeedOverride is True:
    ws["F1"].value = "Major Upd. (est)"
else:
    ws["F1"].value = "Major Upd. (true)"
ws["G1"].value = "Del. Votes"
ws["H1"].value = "Del. Endos"
if not minimize:
    ws["I1"].value = "Embassies"
ws["J1"].value = "WFE"

ws["L1"].value = "World "
ws["M1"].value = "Data"
ws["L2"].value = "Nations"
ws["L3"].value = "Last Major"
ws["L4"].value = "Secs/Nation"
ws["L5"].value = "Nations/Sec"
ws["L6"].value = "Last Minor"
ws["L7"].value = "Secs/Nation"
ws["L8"].value = "Nations/Sec"
ws["L10"].value = "Spyglass Version"
ws["L11"].value = "Date Generated"
ws["M2"].value = CumuNatList[-1]
ws["M3"].value = major
ws["M4"].value = major / CumuNatList[-1]
ws["M5"].value = 1 / (major / CumuNatList[-1])
ws["M6"].value = minor
ws["M7"].value = minor_per_nation
ws["M8"].value = 1 / minor_per_nation
ws["M10"].value = VERSION
ws["M11"].value = f"{datetime.now().year}-{datetime.now().month}-{datetime.now().day}"

for counter, region in enumerate(regions):
    name = region["name"]
    if region["name"] in pless and region["exec"] is True:
        ws.cell(row=counter + 2, column=1).fill = yellowFill
        ws.cell(row=counter + 2, column=2).fill = yellowFill
        name = f"{region['name']}~"
    if region["name"] in fless and region["name"] in pless:
        ws.cell(row=counter + 2, column=1).fill = greenFill
        ws.cell(row=counter + 2, column=2).fill = greenFill
        name = f"{region['name']}~"
    if region["name"] not in pless:
        ws.cell(row=counter + 2, column=1).fill = redFill
        ws.cell(row=counter + 2, column=2).fill = redFill
        name = f"{region['name']}*"

    ws.cell(row=counter + 2, column=1).value = name
    ws.cell(row=counter + 2, column=2).value = region["url"]
    ws.cell(row=counter + 2, column=3).value = region["num_nations"]
    ws.cell(row=counter + 2, column=4).value = region["cumu_nations"]
    ws.cell(row=counter + 2, column=5).value = region["minor"]
    ws.cell(row=counter + 2, column=5).alignment = Alignment(horizontal="right")
    ws.cell(row=counter + 2, column=6).value = region["major"]
    ws.cell(row=counter + 2, column=6).alignment = Alignment(horizontal="right")
    ws.cell(row=counter + 2, column=7).value = region["del_votes"]
    ws.cell(row=counter + 2, column=8).value = region["del_votes"] - 1
    ws.cell(row=counter + 2, column=9).value = (
        ",".join(region["embassies"]) if not minimize else " "
    )
    ws.cell(row=counter + 2, column=10).value = region["wfe"] if not minimize else " "
    ws.cell(row=counter + 2, column=11).value = " "
    if region["del_votes"] == 0:
        ws.cell(row=counter + 2, column=8).fill = redFill

sheet = wb["Spyglass Timesheet"]
sheet.column_dimensions["A"].width = 45
sheet["J1"].alignment = Alignment(horizontal="right")

logger.info("Spreadsheet populated!")
print("Saving spreadsheet...")
wb.save(args.outfile)

logger.info(f"INFO Successfully saved to {args.outfile}")
raise SystemExit(0)